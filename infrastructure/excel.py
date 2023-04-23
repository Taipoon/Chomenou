import os.path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from domain.exporter import IAccountBookExporter
from infrastructure.factories import StatementFactory


class AccountBookExcelExporter(IAccountBookExporter):
    def __init__(self, year: int, output_dir=None):
        super().__init__(year, output_dir=output_dir)
        self._workbook: openpyxl.Workbook or None = None
        self._worksheet: Worksheet or None = None
        self._sheet_name = f"{self.year}年度"

        self._filename = f"{self.year}年度_帳面.xlsx"
        self._filepath: str = os.path.join(os.path.abspath(self.output_dir), self._filename)

        self._statement_repository = StatementFactory.create()

    def export(self):
        # Excelファイルが存在すればそのファイルを開き、なければ新規作成する
        try:
            self._workbook = openpyxl.load_workbook(self._filepath)
        except FileNotFoundError:
            self._workbook = openpyxl.Workbook()
            self._workbook.save(self._filepath)

        # 書き込むシートの名前変更
        self._worksheet: Worksheet = self._workbook.active
        self._worksheet.title = self._sheet_name

        # 1行1列セルに年度を書き込む
        base_pos = {"row": 1, "column": 1}
        self._worksheet.cell(**base_pos, value=f"{self.year}年")

        # 1行目に行ヘッダを書き込む
        rows_header = ["1月", "2月", "3月", "4月", "5月",
                       "6月", "7月", "8月", "9月", "10月",
                       "11月", "12月", "科目別合計"]
        for i, header in enumerate(rows_header, start=1):
            self._worksheet.cell(row=base_pos["row"],
                                 column=base_pos["column"] + i,
                                 value=header)

        # 1列目に列ヘッダを書き込む
        for month_num in range(1, 13):
            data = self._statement_repository.get_monthly_account_summary(year=self.year, month=month_num)
            print(data)

            self._worksheet.cell(row=base_pos["row"] + month_num,
                                 column=base_pos["column"],
                                 value="科目名")

        # 保存して終了
        self._workbook.save(self._filepath)
        self._workbook.close()


if __name__ == '__main__':
    exporter = AccountBookExcelExporter(2025, output_dir=".")
    exporter.export()
